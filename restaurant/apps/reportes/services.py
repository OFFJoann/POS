"""
Servicios de la aplicación reportes.

Contiene la lógica de negocio para generación de reportes
y exportación de datos.
"""
from datetime import date, datetime
from decimal import Decimal
from django.db.models import Sum, Count, Q, F
from django.http import HttpResponse
from django.utils import timezone

from apps.ventas.models import Factura
from apps.productos.models import Producto
from apps.mesas.models import DetallePedido
from apps.caja.models import AperturaCaja, Egreso
from apps.inventario.models import ConsumoInterno


def obtener_datos_dashboard():
    """
    Recolecta todos los datos necesarios para el dashboard.

    Retorna un diccionario con las métricas principales.
    """
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)
    inicio_anio = hoy.replace(month=1, day=1)

    inicio_mes = timezone.make_aware(datetime(inicio_mes.year, inicio_mes.month, inicio_mes.day))
    inicio_anio = timezone.make_aware(datetime(inicio_anio.year, inicio_anio.month, inicio_anio.day))

    facturas_hoy = Factura.objects.filter(created_at__date=hoy)
    ventas_dia = facturas_hoy.aggregate(
        total_ventas=Sum('total'),
        total_efectivo=Sum('total', filter=Q(metodo_pago='efectivo')),
        total_transferencia=Sum('total', filter=Q(metodo_pago='transferencia')),
        descuentos=Sum('descuento'),
    )

    facturas_mes = Factura.objects.filter(created_at__gte=inicio_mes)
    ventas_mes = facturas_mes.aggregate(total=Sum('total'))['total'] or 0

    facturas_anio = Factura.objects.filter(created_at__gte=inicio_anio)
    ventas_anio = facturas_anio.aggregate(total=Sum('total'))['total'] or 0

    productos_hoy = DetallePedido.objects.filter(
        pedido__factura__created_at__date=hoy
    ).aggregate(total=Sum('cantidad'), count=Count('id'))

    mas_vendidos = DetallePedido.objects.values(
        'producto__nombre', 'producto__codigo'
    ).annotate(total_vendido=Sum('cantidad')).order_by('-total_vendido')[:10]

    menos_vendidos = DetallePedido.objects.values(
        'producto__nombre', 'producto__codigo'
    ).annotate(total_vendido=Sum('cantidad')).order_by('total_vendido')[:10]

    egresos_hoy = Egreso.objects.filter(created_at__date=hoy)
    total_egresos = egresos_hoy.aggregate(total=Sum('valor'))['total'] or 0

    egresos_mes = Egreso.objects.filter(created_at__gte=inicio_mes)
    total_egresos_mes = egresos_mes.aggregate(total=Sum('valor'))['total'] or 0

    egresos_anio = Egreso.objects.filter(created_at__gte=inicio_anio)
    total_egresos_anio = egresos_anio.aggregate(total=Sum('valor'))['total'] or 0

    egresos_por_categoria = egresos_hoy.values(
        'categoria__nombre'
    ).annotate(
        total=Sum('valor'),
        cantidad=Count('id')
    ).order_by('-total')

    caja_actual = AperturaCaja.objects.filter(activa=True).first()

    cortesias_hoy = DetallePedido.objects.filter(
        pedido__factura__created_at__date=hoy, es_cortesia=True
    )
    total_cortesias = cortesias_hoy.aggregate(total=Sum('precio_unitario'))['total'] or 0
    cortesias_items = cortesias_hoy.select_related(
        'producto', 'pedido__factura'
    ).order_by('-id')[:20]

    consumos_hoy = ConsumoInterno.objects.filter(created_at__date=hoy)
    consumos_items = consumos_hoy.select_related('producto', 'empleado').order_by('-id')[:20]

    return {
        'ventas_dia': ventas_dia['total_ventas'] or 0,
        'ventas_mes': ventas_mes,
        'ventas_anio': ventas_anio,
        'efectivo_dia': ventas_dia['total_efectivo'] or 0,
        'transferencia_dia': ventas_dia['total_transferencia'] or 0,
        'descuentos_dia': ventas_dia['descuentos'] or 0,
        'total_egresos': total_egresos,
        'total_egresos_mes': total_egresos_mes,
        'total_egresos_anio': total_egresos_anio,
        'egresos_por_categoria': egresos_por_categoria,
        'productos_vendidos': productos_hoy['total'] or 0,
        'mas_vendidos': mas_vendidos,
        'menos_vendidos': menos_vendidos,
        'caja_actual': caja_actual,
        'num_facturas': facturas_hoy.count(),
        'total_cortesias': total_cortesias,
        'cortesias_items': cortesias_items,
        'num_cortesias': cortesias_hoy.count(),
        'consumos_items': consumos_items,
        'num_consumos': consumos_hoy.count(),
    }


def exportar_ventas_excel(facturas):
    """
    Exporta la lista de facturas a un archivo Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['# Factura', 'Fecha', 'Hora', 'Mesa', 'Mesero',
               'Método Pago', 'Subtotal', 'Descuento', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, factura in enumerate(facturas, 2):
        data = [
            factura.numero,
            factura.created_at.strftime('%d/%m/%Y'),
            factura.created_at.strftime('%H:%M'),
            factura.mesa.numero,
            str(factura.mesero),
            factura.get_metodo_pago_display(),
            float(factura.subtotal),
            float(factura.descuento),
            float(factura.total),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col >= 7:
                cell.number_format = '#,##0.00'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
    wb.save(response)
    return response


def exportar_productos_excel(productos):
    """
    Exporta la lista de productos a un archivo Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Código', 'Producto', 'Categoría', 'Precio Venta',
               'Costo', 'Stock Actual', 'Stock Mínimo', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row, prod in enumerate(productos, 2):
        data = [
            prod.codigo, prod.nombre, prod.categoria.nombre,
            float(prod.precio_venta), float(prod.costo),
            float(prod.stock_actual), float(prod.stock_minimo),
            prod.get_estado_display()
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response
